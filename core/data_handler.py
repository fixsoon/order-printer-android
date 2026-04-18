"""Excel 数据读取与处理（处理合并单元格）"""

import openpyxl
import pandas as pd
from datetime import datetime


# Excel 列名（按顺序）
COLUMNS = [
    "订单号", "接龙号", "购买人", "商品名称", "数量",
    "商品金额", "订单总金额", "订单已退款", "订单状态",
    "在线支付金额", "下单时间", "用户备注", "发起人备注",
    "自提点", "收货人", "联系电话", "学校", "学生姓名",
    "年段", "班级", "餐别",
]


def read_excel(filepath):
    """
    读取 Excel 文件，处理接龙号等合并单元格。
    返回 pandas DataFrame。
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # 构建合并单元格映射：左上角值填充到所有合并区域
    merged_map = {}  # (row, col) -> value
    for merged_range in ws.merged_cells.ranges:
        min_row, min_col = merged_range.min_row, merged_range.min_col
        value = ws.cell(min_row, min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_map[(row, col)] = value

    # 读取数据
    rows = []
    # 先从表头找到数据起始行
    header_row = _find_header_row(ws)
    if header_row is None:
        raise ValueError("未找到表头行，请确认 Excel 格式正确")

    max_col = min(ws.max_column, len(COLUMNS))

    for row_idx in range(header_row + 1, ws.max_row + 1):
        row_data = []
        has_data = False
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row_idx, col_idx)
            # 优先使用合并映射的值
            value = merged_map.get((row_idx, col_idx), cell.value)
            if value is not None:
                has_data = True
            row_data.append(value)

        if has_data:
            rows.append(row_data)

    # 构建 DataFrame
    df = pd.DataFrame(rows, columns=COLUMNS[:max_col])

    # 清理数据
    df = _clean_data(df)
    wb.close()
    return df


def _find_header_row(ws):
    """找到包含 '订单号' 或 '接龙号' 的表头行"""
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=False):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if "订单号" in cell.value or "接龙号" in cell.value:
                    return cell.row
    return None


def _clean_data(df):
    """数据清理和类型转换"""
    # 填充接龙号（处理合并单元格残留的空值）
    if "接龙号" in df.columns:
        df["接龙号"] = df["接龙号"].ffill()

    # 数字列转为数值
    for col in ["数量", "商品金额", "订单总金额", "在线支付金额"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # 时间列
    if "下单时间" in df.columns:
        df["下单时间"] = pd.to_datetime(df["下单时间"], errors="coerce")

    # 字符串列去除空白
    str_cols = ["接龙号", "商品名称", "学校", "学生姓名", "年段", "班级",
                "用户备注", "发起人备注", "自提点", "餐别"]
    for col in str_cols:
        if col in df.columns:
            df[col] = df[col].fillna("").astype(str).str.strip()

    return df


def filter_orders(df, chain_start=None, chain_end=None, exclude_cancelled=True):
    """
    筛选订单
    chain_start/end: 接龙号范围（数字）
    exclude_cancelled: 排除取消的订单
    """
    result = df.copy()

    if exclude_cancelled and "订单状态" in result.columns:
        result = result[result["订单状态"] != "订单取消"]

    if chain_start is not None and "接龙号" in result.columns:
        result = result[result["接龙号"].apply(
            lambda x: _extract_chain_num(x) >= chain_start
        )]

    if chain_end is not None and "接龙号" in result.columns:
        result = result[result["接龙号"].apply(
            lambda x: _extract_chain_num(x) <= chain_end
        )]

    return result


def _extract_chain_num(chain_str):
    """从接龙号字符串中提取数字部分"""
    import re
    match = re.search(r'(\d+)', str(chain_str))
    return int(match.group(1)) if match else 0


def sort_by_school_chain(df, schools_order=None):
    """按学校+接龙号排序"""
    if schools_order is None:
        schools_order = ["音西一中", "文光中学", "滨江中学", "二中东校区"]

    df = df.copy()
    df["学校排序"] = df["学校"].apply(
        lambda x: next((i for i, s in enumerate(schools_order) if s in x), 99)
    )
    df = df.sort_values(["学校排序", "接龙号", "下单时间"])
    df = df.drop(columns=["学校排序"], errors="ignore")
    return df


def group_by_chain(df):
    """按接龙号分组，返回字典 {接龙号: DataFrame}"""
    return {chain: group for chain, group in df.groupby("接龙号")}


def group_orders_for_receipt(df):
    """
    将订单按接龙号分组，每组内列出商品明细。
    返回: [{接龙号, 学校, 年段, 班级, 学生姓名, 餐别, 用户备注, 发起人备注, 下单时间, items:[{名称,数量,金额}]}]
    """
    groups = []
    for chain, chain_df in group_by_chain(df):
        # 取第一行的基础信息
        first = chain_df.iloc[0]
        items = []
        for _, row in chain_df.iterrows():
            items.append({
                "name": row.get("商品名称", ""),
                "qty": row.get("数量", 0),
                "price": row.get("商品金额", 0),
            })

        groups.append({
            "chain": chain,
            "school": first.get("学校", ""),
            "grade": first.get("年段", ""),
            "class_name": first.get("班级", ""),
            "student": first.get("学生姓名", ""),
            "meal": first.get("餐别", ""),
            "user_note": first.get("用户备注", ""),
            "admin_note": first.get("发起人备注", ""),
            "time": first.get("下单时间", ""),
            "items": items,
            "total": chain_df["订单总金额"].sum(),
        })

    return groups


def check_school_pickup_mismatch(df, mapping):
    """
    检查学校与自提点不匹配的接龙号
    mapping: {学校名: [允许的自提点列表]}
    返回: [(接龙号, 学校, 自提点)]
    """
    mismatches = []
    seen = set()
    for _, row in df.iterrows():
        chain = row.get("接龙号", "")
        school = row.get("学校", "")
        pickup = row.get("自提点", "")
        if chain in seen:
            continue
        seen.add(chain)
        if school in mapping:
            valid_pickups = mapping[school]
            if pickup not in valid_pickups:
                mismatches.append((chain, school, pickup))
    return mismatches
